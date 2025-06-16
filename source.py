from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

total_sum = 0
count = 0

for row in range(2, max_row + 1):
    product = ws['I' + str(row)].value
    price = ws['K' + str(row)].value

    if isinstance(product, str) and 'LaserJet' in product and isinstance(price, (int, float)):
        total_sum += price
        count += 1

if count > 0:
    average = int(total_sum / count)
else:
    average = 0

print(average)
